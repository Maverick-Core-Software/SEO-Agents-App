from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from seo_agents.crew import OUTPUT_DIR
from seo_agents.website import WEBSITE_ACTION_TYPES, run_website_action, website_adapter_status

# Session 3: imports for lineage/claim fields on action objects
from seo_agents.contracts import ExecutionTask, stable_hash
from seo_agents.evidence import TASK_GRAPH_PATH, write_task_graph


ACTION_QUEUE_FILE = OUTPUT_DIR / "action_queue.json"
ACTION_APPROVALS_FILE = OUTPUT_DIR / "action_approvals.json"
ACTION_COMPLETIONS_FILE = OUTPUT_DIR / "action_completions.json"
ACTION_COMPLETIONS_CONFIG_FILE = Path(os.getenv(
    "ACTION_COMPLETIONS_CONFIG_FILE",
    "config/action-completions.json",
))
ACTION_RUN_DIR = OUTPUT_DIR / "action_runs"
GBP_POSTER_SCRIPT = Path(os.getenv(
    "GBP_POSTER_SCRIPT",
    r"C:\Users\carte\.claude\skills\gbp-poster\driver.mjs",
))
GBP_POSTER_CONFIG = Path(os.getenv(
    "GBP_POSTER_CONFIG",
    r"C:\Users\carte\.codex\plugins\grizzly-gbp-poster\config.local.json",
))
GBP_BROWSER_SESSION_DIR = Path(os.getenv(
    "GBP_BROWSER_SESSION_DIR",
    r"C:\Users\carte\.claude\gbp-session",
))
GBP_POSTER_TIMEOUT_S = int(os.getenv("GBP_POSTER_TIMEOUT_S", "420"))
GBP_POSTER_HEADLESS = os.getenv("GBP_POSTER_HEADLESS", "0").lower() in {"1", "true", "yes", "on"}
GBP_PROFILE_ADAPTER = os.getenv(
    "GBP_PROFILE_ADAPTER",
    r"C:\Workspace\Active\SEO-Agents-App\scripts\gbp-profile-adapter.mjs",
).strip()
GBP_PROFILE_ADAPTER_TIMEOUT_S = int(os.getenv("GBP_PROFILE_ADAPTER_TIMEOUT_S", "300"))
FACEBOOK_POSTER_ADAPTER = os.getenv(
    "FACEBOOK_POSTER_ADAPTER",
    r"C:\Workspace\Active\SEO-Agents-App\scripts\facebook-poster.mjs",
).strip()
# Allow >720s so a slow Veo 3 video render (see gemini-video-generator MAX_POLL_ATTEMPTS)
# plus upload completes before this subprocess is killed.
FACEBOOK_POSTER_TIMEOUT_S = int(os.getenv("FACEBOOK_POSTER_TIMEOUT_S", "900"))
GBP_WORKBOOK_HEADERS = [
    "Date",
    "PostType",
    "Topic",
    "AssetSource",
    "AssetIdOrDescription",
    "CTA",
    "Status",
    "CaptionDraft",
    "ImageLink",
    "Posted",
    "PostedAt",
    "GBPPostUrl",
    "Notes",
]


def _now_iso() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _read_text(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="replace")


def _markdown_body(text: str) -> str:
    stripped = text.strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```[a-zA-Z0-9_-]*\s*", "", stripped)
        stripped = re.sub(r"\s*```\s*$", "", stripped)
    return stripped.strip()


def _load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(exist_ok=True)
    temp_path = path.with_suffix(f"{path.suffix}.tmp")
    temp_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    temp_path.replace(path)


def _run_subprocess(command: list[str], cwd: str, timeout_s: int, display_command: str | None = None) -> dict[str, Any]:
    """Run an adapter subprocess, returning a uniform result record.

    A hung browser run must come back as an adapter failure record (exit 124),
    not an uncaught TimeoutExpired blowing up run_action.
    """
    shown = display_command or " ".join(command)
    try:
        result = subprocess.run(
            command,
            cwd=cwd,
            capture_output=True,
            text=True,
            timeout=timeout_s,
            check=False,
        )
    except subprocess.TimeoutExpired as error:
        return {
            "exit_code": 124,
            "command": shown,
            "stdout": (error.stdout or b"").decode(errors="replace") if isinstance(error.stdout, bytes) else (error.stdout or ""),
            "stderr": f"Adapter timed out after {timeout_s}s.",
        }
    except OSError as error:
        return {
            "exit_code": 126,
            "command": shown,
            "stdout": "",
            "stderr": f"Failed to launch adapter: {error}",
        }
    return {
        "exit_code": result.returncode,
        "command": shown,
        "stdout": result.stdout.strip(),
        "stderr": result.stderr.strip(),
    }


def _last_json_line(stdout: str) -> dict[str, Any]:
    """Adapters emit a single-line JSON result as their final stdout line."""
    for line in reversed(stdout.splitlines()):
        line = line.strip()
        if line.startswith("{") and line.endswith("}"):
            try:
                return json.loads(line)
            except json.JSONDecodeError:
                continue
    return {}


def _extract_numbered_field(block: str, label: str) -> str:
    match = re.search(rf"\d+\)\s*\*\*{re.escape(label)}\*\*:\s*(.+)", block)
    return match.group(1).strip() if match else ""


def _extract_list_after(block: str, label: str) -> list[str]:
    marker = re.search(rf"\d+\)\s*\*\*{re.escape(label)}\*\*:\s*", block)
    if not marker:
        return []
    tail = block[marker.end():]
    next_field = re.search(r"^\d+\)\s*\*\*.+?\*\*:", tail, flags=re.MULTILINE)
    section = tail[:next_field.start()] if next_field else tail
    return [
        line.strip()[2:].strip()
        for line in section.splitlines()
        if line.strip().startswith("- ")
    ]


def _extract_completion_blocks(text: str) -> dict[str, dict[str, str]]:
    body = _markdown_body(text)
    parts = re.split(r"(?=^### COMPLETION REPORT)", body, flags=re.MULTILINE)
    completions: dict[str, dict[str, str]] = {}
    for part in parts:
        task_id = re.search(r"^Task ID:\s*([A-Z]+-?\d+)", part, flags=re.MULTILINE)
        if not task_id:
            continue
        completion: dict[str, str] = {"task_id": task_id.group(1)}
        for label, key in (
            ("Task Title", "title"),
            ("Assigned Agent", "agent"),
            ("Status", "completion_status"),
            ("Action Taken", "action_taken"),
            ("Deliverable Location", "deliverable_location"),
            ("Definition of Done Met", "definition_of_done"),
            ("If Partial or Blocked", "blocker"),
            ("Owner Sign-Off Needed", "owner_signoff_needed"),
        ):
            match = re.search(rf"^{re.escape(label)}:\s*(.+)", part, flags=re.MULTILINE)
            if match:
                completion[key] = match.group(1).strip()
        completions[completion["task_id"]] = completion
    task_parts = re.split(r"(?=^## Task \d+:)", body, flags=re.MULTILINE)
    for part in task_parts:
        task_id = re.search(r"^### Task ID:\s*([A-Z]+-?\d+)", part, flags=re.MULTILINE)
        if not task_id:
            continue
        completion = {"task_id": task_id.group(1)}
        title = re.search(r"^## Task \d+:\s*(.+)", part, flags=re.MULTILINE)
        if title:
            completion["title"] = title.group(1).strip()
        for label, key in (
            ("Status", "completion_status"),
            ("Definition of Done Met", "definition_of_done"),
            ("Deliverable Location", "deliverable_location"),
            ("If Partial or Blocked", "blocker"),
            ("Owner Sign-Off Needed", "owner_signoff_needed"),
        ):
            match = re.search(rf"^### {re.escape(label)}:\s*(.+)", part, flags=re.MULTILINE)
            if match:
                completion[key] = match.group(1).strip()
        action_match = re.search(
            r"^### Action Taken:\s*\n(.*?)(?=^### |\n---|\Z)",
            part,
            flags=re.MULTILINE | re.DOTALL,
        )
        if action_match:
            completion["action_taken"] = " ".join(action_match.group(1).split())
        completions[completion["task_id"]] = completion
    return completions


def _completions_from_json(payload: Any) -> dict[str, dict[str, str]]:
    """Parse a structured completion report (crew output_json schema)."""
    entries = payload.get("completions", []) if isinstance(payload, dict) else payload
    if not isinstance(entries, list):
        return {}
    completions: dict[str, dict[str, str]] = {}
    for entry in entries:
        if not isinstance(entry, dict) or not entry.get("task_id"):
            continue
        completions[str(entry["task_id"]).strip()] = {
            key: str(value).strip()
            for key, value in entry.items()
            if value is not None
        }
    return completions


def _load_completions() -> dict[str, dict[str, str]]:
    completions: dict[str, dict[str, str]] = {}
    for stem in ("content_completion", "assets_completion", "technical_completion", "website_completion"):
        # Structured JSON (preferred, regex-free) wins; markdown is the fallback
        # for older runs or models that failed structured output. A stale JSON
        # from a previous run must not shadow a fresher markdown report.
        json_path = OUTPUT_DIR / f"{stem}.json"
        md_path = OUTPUT_DIR / f"{stem}.md"
        if json_path.exists() and (not md_path.exists() or json_path.stat().st_mtime >= md_path.stat().st_mtime):
            parsed = _completions_from_json(_load_json(json_path, {}))
            if parsed:
                completions.update(parsed)
                continue
        completions.update(_extract_completion_blocks(_read_text(OUTPUT_DIR / f"{stem}.md")))
    return completions


def _load_action_completions() -> dict[str, dict[str, Any]]:
    """Load owner-verified / manual completion overrides.

    Keys are action IDs (e.g. ``task-t002``). Each entry must have a ``status``
    field (e.g. ``"verified"``) and optionally ``verified_at``, ``verified_by``,
    and ``details``.
    """
    completions = _load_json(ACTION_COMPLETIONS_CONFIG_FILE, {})
    completions.update(_load_json(ACTION_COMPLETIONS_FILE, {}))
    return completions


_WEBSITE_TYPE_KEYWORDS: list[tuple[str, tuple[str, ...]]] = [
    ("website_blog_post", ("blog",)),
    ("website_contact_form_update", ("contact form", "form field", "form submission")),
    ("website_gallery_update", ("gallery", "project photo", "before and after", "before/after")),
    ("website_hours_update", ("hours update", "business hours", "opening hours", "hours of operation")),
    ("website_faq_update", ("faq", "frequently asked")),
    ("website_service_page_update", ("service page", "services section", "service card", "service area")),
    ("website_layout_update", ("layout", "navigation", "nav link", "header", "footer", "hero")),
]


def _infer_action_type(executor: str, title: str, steps: list[str]) -> str:
    haystack = f"{executor} {title} {' '.join(steps)}".lower()
    if "gbp" in haystack or "google business" in haystack:
        return "gbp_profile_update"
    if "facebook" in haystack or "fb" in executor.lower():
        return "publish_facebook_post"
    if "review" in haystack and "website" not in executor.lower():
        return "review_management"
    for action_type, keywords in _WEBSITE_TYPE_KEYWORDS:
        if any(keyword in haystack for keyword in keywords):
            return action_type
    if "website" in haystack or "landing page" in haystack or "copy" in haystack or "content" in executor.lower():
        return "website_copy_update"
    return "manual_followup"


def _risk_for_action(action_type: str) -> str:
    if action_type in {"website_layout_update", "website_contact_form_update", "gbp_profile_update"}:
        return "high"
    if action_type in WEBSITE_ACTION_TYPES or action_type in {"review_management", "publish_gbp_post", "publish_facebook_post"}:
        return "medium"
    return "low"


def _platform_for_action(action_type: str) -> str:
    if action_type in WEBSITE_ACTION_TYPES:
        return "website"
    return {
        "gbp_profile_update": "google_business_profile",
        "publish_gbp_post": "google_business_profile",
        "publish_facebook_post": "facebook_page",
        "review_management": "review_platforms",
    }.get(action_type, "manual")


def _status_for_action(
    completion: dict[str, str],
    dependencies: list[str],
    completion_override: dict[str, Any] | None = None,
) -> str:
    # Owner-verified / manual completion overrides short-circuit all other logic.
    if completion_override and completion_override.get("status") == "verified":
        return "verified"
    status = completion.get("completion_status", "").upper()
    dependency_text = " ".join(dependencies).lower()
    blocker_text = completion.get("blocker", "").lower()
    if status == "COMPLETE" and completion.get("definition_of_done", "").upper() == "YES":
        return "dry_run_ready"
    if "access" in dependency_text or "access" in blocker_text or status == "BLOCKED":
        return "blocked_access"
    if completion.get("owner_signoff_needed", "").upper() == "YES":
        return "needs_approval"
    if status in {"COMPLETE", "PARTIAL"}:
        return "dry_run_ready"
    return "needs_review"


def _priority_for_action(action_type: str, risk: str, executor: str) -> dict[str, Any]:
    """Session 3: derive priority tier from action characteristics (v1 proposed defaults)."""
    # Impact heuristic: lead-gen actions > content > maintenance
    impact = {
        "gbp_profile_update": 0.85,
        "website_contact_form_update": 0.80,
        "website_layout_update": 0.70,
        "website_blog_post": 0.65,
        "website_service_page_update": 0.60,
        "website_gallery_update": 0.55,
        "website_hours_update": 0.50,
        "website_faq_update": 0.50,
        "website_copy_update": 0.50,
    }.get(action_type, 0.40)

    # Confidence: high-risk adapter actions get higher confidence (we have tooling)
    confidence = 0.7 if risk == "high" else 0.5 if risk == "medium" else 0.4

    # Urgency: blog posts are time-sensitive ("this week")
    urgency = 0.8 if action_type == "website_blog_post" else 0.5

    # Strategic alignment: content updates and local fixes are core SEO
    strategic_alignment = 0.7 if action_type in {
        "gbp_profile_update", "website_blog_post", "website_service_page_update",
        "website_copy_update", "website_faq_update", "website_hours_update",
    } else 0.5

    score = round(0.35 * impact + 0.30 * confidence + 0.20 * urgency + 0.15 * strategic_alignment, 2)
    # Map score + confidence to tier
    if score >= 0.80 and confidence >= 0.6:
        tier = "P0"
    elif score >= 0.60 and confidence >= 0.5:
        tier = "P1"
    elif score >= 0.40:
        tier = "P2"
    else:
        tier = "P3"
    return {"tier": tier, "score": score, "formula_version": "priority-v1"}


def _confidence_for_action(action_type: str, risk: str) -> dict[str, Any]:
    """Session 3: derive task confidence label."""
    conf = 0.5 if risk == "medium" else 0.7 if risk == "high" else 0.4
    label = "high" if conf >= 0.6 else "medium" if conf >= 0.4 else "low"
    return {"label": label, "score": round(conf, 2)}


def _approval_class_for_action(action_type: str) -> str:
    """Session 3: determine approval class based on risk."""
    if action_type in {"gbp_profile_update", "website_layout_update", "website_contact_form_update"}:
        return "mandatory"
    if action_type in WEBSITE_ACTION_TYPES or action_type in {"review_management", "publish_gbp_post", "publish_facebook_post"}:
        return "sampled"
    return "none"


def _uncertainty_for_action(action_type: str) -> dict[str, Any]:
    """Session 3: uncertainty metadata."""
    gaps: dict[str, Any] = {"proxy_metrics_used": [], "gap_reason": None, "blocked_by": []}
    if action_type == "website_blog_post":
        gaps["proxy_metrics_used"] = ["content_topic_trend"]
        gaps["gap_reason"] = "No live traffic data to validate topic selection"
    return gaps


def parse_execution_actions(run_id: str = "") -> list[dict[str, Any]]:
    queue_text = _markdown_body(_read_text(OUTPUT_DIR / "grizzly_execution_queue.md"))
    completions = _load_completions()
    completion_overrides = _load_action_completions()
    parts = re.split(r"(?=^## Task \d+:)", queue_text, flags=re.MULTILINE)
    actions: list[dict[str, Any]] = []
    for part in parts:
        task_id = _extract_numbered_field(part, "Task ID")
        if not task_id:
            continue
        title = _extract_numbered_field(part, "Task Title")
        executor = _extract_numbered_field(part, "Assigned Execution Agent")
        steps = _extract_list_after(part, "Exact Action Steps")
        dependencies = _extract_list_after(part, "Dependencies / Required Inputs")
        verification = _extract_list_after(part, "Verification Checklist")
        completion = completions.get(task_id, {})
        action_id = f"task-{task_id.lower()}"
        override = completion_overrides.get(action_id)
        action_type = _infer_action_type(executor, title, steps)
        platform = _platform_for_action(action_type)
        status = _status_for_action(completion, dependencies, override)
        risk = _risk_for_action(action_type)
        priority = _priority_for_action(action_type, risk, executor)
        confidence = _confidence_for_action(action_type, risk)
        approval_class = _approval_class_for_action(action_type)
        uncertainty = _uncertainty_for_action(action_type)
        # Build deterministic idempotency key from action attributes
        idem_seed = f"{action_type}:{title}:{executor}:{'|'.join(steps)}"
        idempotency_key = stable_hash(prefix="idem_", data=idem_seed)
        # Build verification dict from checklist
        verification_dict: dict[str, Any] = {}
        if verification:
            verification_dict["checklist"] = verification
        # Build rollback string
        rollback = f"Revert: revert {action_type} changes for '{title}'"
        # Build supporting_claim_ids (extract from queue text if present; empty if not)
        claim_match = re.search(r"\*\*Supporting Claim IDs\*\*:\s*(.+)", part)
        supporting_claim_ids: list[str] = []
        if claim_match:
            supporting_claim_ids = [
                cid.strip()
                for cid in claim_match.group(1).split(",")
                if cid.strip().startswith("claim_")
            ]
        # Build preconditions
        preconditions: list[str] = []
        if platform == "website" and not action_type.startswith("website_blog_post"):
            preconditions.append("Website repo must be cloned and accessible")
        if action_type in {"gbp_profile_update", "publish_gbp_post"}:
            preconditions.append("Google Business Profile access required")
        if action_type == "publish_facebook_post":
            preconditions.append("Facebook Page access token required")
        # Build acceptance criteria from definition of done
        acceptance_criteria: list[str] = []
        if completion:
            dod = completion.get("definition_of_done", "")
            if dod:
                acceptance_criteria.append(dod)
        acceptance_criteria.append(f"{action_type} action verified complete")
        actions.append({
            "id": action_id,
            "source": "execution_queue",
            "source_task_id": task_id,
            "title": title,
            "assigned_agent": executor,
            "action_type": action_type,
            "platform": platform,
            "risk": risk,
            "status": status,
            "priority": priority,
            "due_window": _extract_numbered_field(part, "Due Window"),
            "steps": steps,
            "dependencies": dependencies,
            "verification_checklist": verification,
            "completion": completion,
            "completion_override": override,
            "approval_required": status != "verified" and (
                completion.get("owner_signoff_needed", "").upper() == "YES" or bool(dependencies)
            ),
            "live_adapter": "website_manager" if platform == "website" else None,
            # Session 3 additive lineage fields
            "supporting_claim_ids": supporting_claim_ids,
            "confidence": confidence,
            "approval_class": approval_class,
            "uncertainty": uncertainty,
            "idempotency_key": idempotency_key,
            "verification": verification_dict,
            "rollback": rollback,
            "preconditions": preconditions,
            "acceptance_criteria": acceptance_criteria,
        })
    return actions


def _parse_gbp_post_blocks(text: str) -> list[dict[str, str]]:
    blocks = re.split(r"^\s*---\s*$", text, flags=re.MULTILINE)
    posts: list[dict[str, str]] = []
    for block in blocks:
        cleaned = block.replace("**", "")
        if "DAY:" not in cleaned or "HEADLINE:" not in cleaned:
            continue
        post: dict[str, str] = {}
        for label in ("DAY", "DATE", "SERVICE", "TOPIC", "TREND_TIE", "HEADLINE", "BODY", "CAPTION", "PHOTO_FILE", "CTA", "STATUS"):
            match = re.search(rf"^{label}:\s*(.+)", cleaned, flags=re.MULTILINE)
            if match:
                post[label.lower()] = match.group(1).strip()
        if post:
            posts.append(post)
    return posts


def parse_gbp_post_actions(run_id: str = "") -> list[dict[str, Any]]:
    posts = _parse_gbp_post_blocks(_read_text(OUTPUT_DIR / "gbp_posting_schedule.md"))
    actions: list[dict[str, Any]] = []
    for index, post in enumerate(posts, start=1):
        post_id = post.get("date") or f"day-{index}"
        action_id = f"gbp-post-{post_id}"
        action_type = "publish_gbp_post"
        idem_seed = f"{action_type}:{post.get('headline', '')}:{post_id}"
        actions.append({
            "id": action_id,
            "source": "gbp_posting_schedule",
            "source_task_id": f"GBP-{index:03d}",
            "title": post.get("headline") or f"GBP post day {index}",
            "assigned_agent": "Grizzly GBP Poster Agent",
            "action_type": action_type,
            "platform": "google_business_profile",
            "risk": "medium",
            "status": "needs_approval" if "approval" in post.get("status", "").lower() else "dry_run_ready",
            "priority": {"tier": "P1", "score": 0.60, "formula_version": "priority-v1"},
            "due_window": post.get("date") or "",
            "steps": [
                "Review post copy and photo selection.",
                "Publish to Google Business Profile after approval.",
            ],
            "dependencies": ["Owner approval", "Google Business Profile access"],
            "verification_checklist": [
                "Confirm post is visible on Google Business Profile.",
                "Confirm selected photo was used.",
                "Archive or mark photo as used after publishing.",
            ],
            "post": post,
            "approval_required": True,
            "live_adapter": "google_business_profile",
            # Session 3 additive lineage fields
            "supporting_claim_ids": [],
            "confidence": {"label": "medium", "score": 0.5},
            "approval_class": "mandatory",
            "uncertainty": {"proxy_metrics_used": [], "gap_reason": None, "blocked_by": []},
            "idempotency_key": stable_hash(prefix="idem_", data=idem_seed),
            "verification": {"checklist": [
                "Confirm post is visible on Google Business Profile.",
                "Confirm selected photo was used.",
                "Archive or mark photo as used after publishing.",
            ]},
            "rollback": "Unpublish or delete the GBP post from the profile.",
            "preconditions": ["Google Business Profile access", "Owner approval"],
            "acceptance_criteria": ["Post visible on GBP profile with selected photo"],
        })
    return actions


def _parse_facebook_post_blocks(text: str) -> list[dict[str, str]]:
    blocks = re.split(r"^\s*---\s*$", text, flags=re.MULTILINE)
    posts: list[dict[str, str]] = []
    for block in blocks:
        cleaned = block.replace("**", "")
        if "DAY:" not in cleaned or "HOOK:" not in cleaned:
            continue
        post: dict[str, str] = {}
        for label in ("DAY", "DATE", "TYPE", "SERVICE", "HOOK", "BODY", "CTA", "HASHTAGS", "PHOTO_FILE", "VIDEO_PROMPT", "STATUS"):
            match = re.search(rf"^{label}:\s*(.+)", cleaned, flags=re.MULTILINE)
            if match:
                post[label.lower()] = match.group(1).strip()
        if post:
            posts.append(post)
    return posts


def parse_facebook_post_actions(run_id: str = "") -> list[dict[str, Any]]:
    posts = _parse_facebook_post_blocks(_read_text(OUTPUT_DIR / "facebook_posting_schedule.md"))
    actions: list[dict[str, Any]] = []
    for index, post in enumerate(posts, start=1):
        post_id = post.get("date") or f"day-{index}"
        action_id = f"fb-post-{post_id}"
        action_type = "publish_facebook_post"
        idem_seed = f"{action_type}:{post.get('hook', '')}:{post_id}"
        actions.append({
            "id": action_id,
            "source": "facebook_posting_schedule",
            "source_task_id": f"FB-{index:03d}",
            "title": post.get("hook") or post.get("body", "")[:60] or f"Facebook post day {index}",
            "assigned_agent": "Grizzly Facebook Poster Agent",
            "action_type": action_type,
            "platform": "facebook_page",
            "risk": "medium",
            "status": "needs_approval" if "approval" in post.get("status", "").lower() else "dry_run_ready",
            "priority": {"tier": "P1", "score": 0.60, "formula_version": "priority-v1"},
            "due_window": post.get("date") or "",
            "steps": [
                "Review post hook, body, and hashtags.",
                "For video posts: verify Gemini video prompt before publishing.",
                "Publish to Facebook Page after approval.",
            ],
            "dependencies": ["Owner approval", "Facebook Page Access Token"],
            "verification_checklist": [
                "Confirm post is visible on Facebook Business Page.",
                "For video posts: confirm video rendered correctly.",
            ],
            "post": post,
            "approval_required": True,
            "live_adapter": "facebook_page",
            # Session 3 additive lineage fields
            "supporting_claim_ids": [],
            "confidence": {"label": "medium", "score": 0.5},
            "approval_class": "mandatory",
            "uncertainty": {"proxy_metrics_used": [], "gap_reason": None, "blocked_by": []},
            "idempotency_key": stable_hash(prefix="idem_", data=idem_seed),
            "verification": {"checklist": [
                "Confirm post is visible on Facebook Business Page.",
                "For video posts: confirm video rendered correctly.",
            ]},
            "rollback": "Delete the Facebook post from the Business Page.",
            "preconditions": ["Facebook Page access token", "Owner approval"],
            "acceptance_criteria": ["Post visible on Facebook Business Page"],
        })
    return actions


def _apply_approvals(actions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    approvals = _load_json(ACTION_APPROVALS_FILE, {})
    for action in actions:
        approval = approvals.get(action["id"])
        if not approval:
            continue
        action["approval"] = approval
        if action["status"] == "needs_approval":
            action["status"] = "approved"
    return actions


def _load_latest_runs() -> dict[str, dict[str, Any]]:
    """Latest run record per action id, so executed/failed state survives queue rebuilds."""
    latest: dict[str, dict[str, Any]] = {}
    if not ACTION_RUN_DIR.exists():
        return latest
    for run_file in sorted(ACTION_RUN_DIR.glob("run-*.json")):
        record = _load_json(run_file, None)
        if not isinstance(record, dict) or not record.get("action_id"):
            continue
        latest[record["action_id"]] = record
    return latest


_RUN_STATUS_TO_ACTION_STATUS = {
    "live_complete": "executed",
    "live_unverified": "needs_verification",
    "adapter_failed": "failed",
}


def _apply_run_results(actions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    runs = _load_latest_runs()
    for action in actions:
        record = runs.get(action["id"])
        if not record:
            continue
        action["last_run"] = {
            "id": record.get("id"),
            "live": record.get("live"),
            "status": record.get("status"),
            "message": record.get("message"),
            "created_at": record.get("created_at"),
        }
        # Owner verification still outranks everything; otherwise reflect reality.
        if action["status"] != "verified":
            mapped = _RUN_STATUS_TO_ACTION_STATUS.get(record.get("status", ""))
            if mapped:
                action["status"] = mapped
    return actions


def build_action_queue(run_id: str = "") -> dict[str, Any]:
    actions = _apply_run_results(_apply_approvals([
        *parse_execution_actions(run_id),
        *parse_gbp_post_actions(run_id),
        *parse_facebook_post_actions(run_id),
    ]))
    adapters = {
        "website_manager": website_adapter_status(),
        "google_business_profile": gbp_adapter_status(),
        "facebook_page": facebook_adapter_status(),
    }
    summary = {
        "total": len(actions),
        "needs_approval": sum(1 for action in actions if action["status"] == "needs_approval"),
        "approved": sum(1 for action in actions if action["status"] == "approved"),
        "verified": sum(1 for action in actions if action["status"] == "verified"),
        "blocked_access": sum(1 for action in actions if action["status"] == "blocked_access"),
        "dry_run_ready": sum(1 for action in actions if action["status"] == "dry_run_ready"),
        "executed": sum(1 for action in actions if action["status"] == "executed"),
        "failed": sum(1 for action in actions if action["status"] == "failed"),
        "needs_verification": sum(1 for action in actions if action["status"] == "needs_verification"),
        "high_risk": sum(1 for action in actions if action["risk"] == "high"),
    }
    return {
        "version": "1.0.0",
        "generated_at": _now_iso(),
        "workflow_id": "grizzly-seo",
        "adapters": adapters,
        "summary": summary,
        "actions": actions,
    }


def _detect_dependency_cycles(actions: list[dict[str, Any]]) -> list[str]:
    """Session 3: detect cycles in the task dependency graph. Returns list of cycle task_ids."""
    # Build adjacency map
    task_deps: dict[str, list[str]] = {}
    action_ids: set[str] = set()
    for action in actions:
        aid = action.get("id", action.get("source_task_id", ""))
        if not aid:
            continue
        task_deps[aid] = action.get("dependencies", [])
        action_ids.add(aid)

    # DFS cycle detection
    visited: set[str] = set()
    rec_stack: set[str] = set()
    cycles: list[str] = []

    def dfs(node: str) -> bool:
        visited.add(node)
        rec_stack.add(node)
        for dep in task_deps.get(node, []):
            if dep not in visited:
                if dep in action_ids and dfs(dep):
                    return True
            elif dep in rec_stack:
                cycles.append(node)
                return True
        rec_stack.discard(node)
        visited.add(node)
        return False

    for task_id in task_deps:
        if task_id not in visited:
            dfs(task_id)
    return cycles


def _unresolved_contradiction_ids(evidence_path: Path = None) -> list[str]:
    """Session 3: extract claim IDs that have unresolved material contradictions."""
    # Check evidence_package.json for unresolved contradictions
    if evidence_path is None:
        from seo_agents.evidence import EVIDENCE_PACKAGE_PATH
        evidence_path = EVIDENCE_PACKAGE_PATH
    if not evidence_path.exists():
        return []
    try:
        raw = json.loads(evidence_path.read_text(encoding="utf-8"))
        evidence_list = raw.get("evidence", [])
        result = validate_evidence_package(evidence_list)
        # Find claims with unresolved contradictions
        unres: list[str] = []
        for g in result.get("gates", []):
            if g.get("gate") == "unresolved_contradiction":
                cid = g.get("claim_id", "")
                if cid and cid not in unres:
                    unres.append(cid)
        return unres
    except Exception:
        return []


def write_action_queue(run_id: str = "") -> dict[str, Any]:
    payload = build_action_queue(run_id)
    _write_json(ACTION_QUEUE_FILE, payload)
    # Session 3: write task_graph.json with lineage fields and dependency validation
    _write_task_graph_from_actions(payload.get("actions", []), run_id)
    return payload


def _write_task_graph_from_actions(actions: list[dict[str, Any]], run_id: str) -> None:
    """Session 3: convert actions to task objects and write task_graph.json."""
    # Detect cycles and unresolved contradictions
    cycles = _detect_dependency_cycles(actions)
    contradict_claims = _unresolved_contradiction_ids()

    # Build task list from actions
    tasks: list[dict[str, Any]] = []
    for action in actions:
        # Check if this task is blocked by unresolved contradictions
        source_claim_ids = action.get("supporting_claim_ids", [])
        blocked_by_contradiction = False
        for claim_id in source_claim_ids:
            if claim_id in contradict_claims:
                blocked_by_contradiction = True
                break

        # Build status
        if blocked_by_contradiction:
            status = "blocked"
        elif action.get("status") in {"needs_approval", "blocked_access"}:
            status = "waiting_on_owner" if action["status"] == "needs_approval" else "waiting_on_tool_access"
        elif action.get("status") == "dry_run_ready":
            status = "ready"
        elif action.get("status") == "verified":
            status = "verified"
        else:
            status = "ready"

        # Build dependencies list
        deps = action.get("dependencies", [])
        # Add cycle-blocked tasks
        for cycle_task in cycles:
            if cycle_task not in deps:
                deps.append(cycle_task)

        # Build task object
        source_task_id = action.get("source_task_id", "")
        task = {
            "task_id": f"T-{run_id}-{action.get('source_task_id', '?').replace('-','_')[:10]}",
            "run_id": run_id,
            "title": action.get("title", ""),
            "task_type": action.get("action_type", "content_update"),
            "supporting_claim_ids": source_claim_ids,
            "owner": _owner_for_action(action.get("action_type", ""), action.get("assigned_agent", "")),
            "priority": action.get("priority", {"tier": "P2", "score": 0.4, "formula_version": "priority-v1"}),
            "confidence": action.get("confidence", {"label": "medium", "score": 0.5}),
            "dependencies": deps,
            "preconditions": action.get("preconditions", []),
            "acceptance_criteria": action.get("acceptance_criteria", []),
            "verification": action.get("verification", {}),
            "rollback": action.get("rollback", ""),
            "approval_class": action.get("approval_class", "none"),
            "uncertainty": action.get("uncertainty", {"proxy_metrics_used": [], "gap_reason": None, "blocked_by": []}),
            "idempotency_key": action.get("idempotency_key", ""),
            "status": status,
        }
        tasks.append(task)

    # Session 3: create research_gap tasks for unresolved claim contradictions
    for claim_id in contradict_claims:
        task = {
            "task_id": f"T-{run_id}-RG-{claim_id[-8:]:0>8}",
            "run_id": run_id,
            "title": f"Resolve contradiction for claim {claim_id}",
            "task_type": "research_gap",
            "supporting_claim_ids": [claim_id],
            "owner": "owner_review",
            "priority": {"tier": "P0", "score": 0.9, "formula_version": "priority-v1"},
            "confidence": {"label": "high", "score": 0.8},
            "dependencies": [],
            "preconditions": ["Review evidence_package.json and claim_graph.json"],
            "acceptance_criteria": ["Contradiction resolved or evidence re-verified"],
            "verification": {"checklist": ["Check that claim is confirmed or rejected with evidence"]},
            "rollback": "No rollback needed — this is a research task",
            "approval_class": "mandatory",
            "uncertainty": {"proxy_metrics_used": ["contradiction_density"], "gap_reason": None, "blocked_by": [claim_id]},
            "idempotency_key": stable_hash(prefix="idem_", data=f"research_gap:{claim_id}"),
            "status": "research_gap",
        }
        tasks.append(task)

    # Write task graph
    write_task_graph(tasks, run_id)


def _owner_for_action(action_type: str, assigned_agent: str) -> str:
    """Map action type / executor to owner field."""
    if "website" in action_type or "Website" in assigned_agent or "Local Content" in assigned_agent:
        return "content_executor"
    if "gbp" in action_type or "GBP" in action_type or "Local Presence" in assigned_agent:
        return "local_presence_assets"
    if "technical" in action_type or "Technical" in assigned_agent:
        return "content_executor"  # technical goes to content_executor for now
    if "review" in action_type or "Review" in assigned_agent:
        return "local_presence_assets"
    return "website_manager"


def approve_action(action_id: str, approved_by: str = "owner", note: str = "") -> dict[str, Any]:
    queue = write_action_queue()
    action = next((item for item in queue["actions"] if item["id"] == action_id), None)
    if not action:
        raise ValueError(f"Unknown action id: {action_id}")
    approvals = _load_json(ACTION_APPROVALS_FILE, {})
    approvals[action_id] = {
        "approved_by": approved_by,
        "approved_at": _now_iso(),
        "note": note,
    }
    _write_json(ACTION_APPROVALS_FILE, approvals)
    if action.get("live_adapter") == "google_business_profile":
        sync_gbp_schedule_to_workbook(dry_run=False)
        _mark_gbp_workbook_status(action, "Approved")
    return write_action_queue()


def run_action(action_id: str, live: bool = False) -> dict[str, Any]:
    queue = write_action_queue()
    action = next((item for item in queue["actions"] if item["id"] == action_id), None)
    if not action:
        raise ValueError(f"Unknown action id: {action_id}")
    if live and action.get("approval_required") and not action.get("approval"):
        result_status = "blocked_approval"
        message = "Live execution requires approval first."
        command_result = None
    elif action.get("action_type") == "gbp_profile_update":
        command_result = _run_gbp_profile_adapter(action, live=live)
        driver_result = _last_json_line(command_result.get("stdout", ""))
        command_result["driver_result"] = driver_result or None
        if command_result["exit_code"] == 0 and not live:
            result_status = "dry_run_complete"
            message = "GBP profile update dry run completed."
        elif command_result["exit_code"] == 0:
            result_status = "live_complete"
            message = f"GBP profile updated: {', '.join(driver_result.get('updates_requested', []))}"
        else:
            result_status = "adapter_failed"
            message = "GBP profile adapter failed."
    elif action.get("live_adapter") == "google_business_profile":
        command_result = _run_gbp_poster(action, live=live)
        driver_result = _last_json_line(command_result.get("stdout", ""))
        command_result["driver_result"] = driver_result or None
        if command_result["exit_code"] == 0 and not live:
            result_status = "dry_run_complete"
            message = "GBP poster dry run completed."
        elif command_result["exit_code"] == 0 and driver_result.get("verified"):
            result_status = "live_complete"
            message = "GBP post published and verified in the Posts list."
            try:
                _mark_gbp_workbook_posted(action, post_url=driver_result.get("postUrl"))
            except Exception as error:
                # The post IS live; never lose that fact over a workbook write issue.
                message = f"GBP post published and verified, but the workbook could not be updated: {error}"
        elif command_result["exit_code"] == 3 or (command_result["exit_code"] == 0 and driver_result.get("result") == "posted"):
            # Submitted but unverified: do NOT mark posted and do NOT auto-retry —
            # a blind retry can publish a duplicate.
            result_status = "live_unverified"
            message = "GBP post was submitted but could not be verified. Check the profile manually before retrying."
        else:
            result_status = "adapter_failed"
            message = "GBP poster adapter failed."
    elif action.get("live_adapter") == "facebook_page":
        command_result = _run_facebook_poster(action, live=live)
        driver_result = _last_json_line(command_result.get("stdout", ""))
        command_result["driver_result"] = driver_result or None
        if command_result["exit_code"] == 0 and not live:
            result_status = "dry_run_complete"
            message = "Facebook poster dry run completed."
        elif command_result["exit_code"] == 0 and driver_result.get("status") == "success":
            result_status = "live_complete"
            message = f"Facebook post published. Post ID: {driver_result.get('post_id')}"
        else:
            result_status = "adapter_failed"
            message = f"Facebook poster adapter failed: {driver_result.get('message', command_result.get('stderr', ''))}"
    elif action.get("live_adapter") == "website_manager":
        command_result = run_website_action(action, live=live)
        adapter_status_value = command_result.get("status", "")
        if adapter_status_value == "pushed":
            result_status = "live_complete"
            message = f"Website change committed and pushed ({command_result.get('commit')}). Vercel deploy triggered."
        elif adapter_status_value == "preview":
            result_status = "dry_run_complete"
            message = f"Website edit preview written to {command_result.get('preview_dir')}. No live change made."
        elif adapter_status_value == "push_failed":
            result_status = "live_unverified"
            message = "Website change committed locally but git push failed. Push manually, then verify the deploy."
        else:
            result_status = "adapter_failed"
            message = command_result.get("message") or "Website adapter failed."
    elif live:
        result_status = "blocked_adapter"
        message = f"No live adapter configured for {action['platform']} yet."
        command_result = None
    else:
        result_status = "dry_run_complete"
        message = "Dry run generated execution payload only. No live system was changed."
        command_result = None

    run_record = {
        "id": f"run-{datetime.now(UTC).strftime('%Y%m%d%H%M%S')}-{action_id}",
        "action_id": action_id,
        "live": live,
        "status": result_status,
        "message": message,
        "action": action,
        "command_result": command_result,
        "created_at": _now_iso(),
    }
    ACTION_RUN_DIR.mkdir(exist_ok=True)
    _write_json(ACTION_RUN_DIR / f"{run_record['id']}.json", run_record)
    return run_record





def _run_gbp_poster(action: dict[str, Any], live: bool) -> dict[str, Any]:
    if not GBP_POSTER_SCRIPT.exists():
        return {
            "exit_code": 127,
            "command": str(GBP_POSTER_SCRIPT),
            "stdout": "",
            "stderr": f"GBP poster script not found: {GBP_POSTER_SCRIPT}",
        }
    date_value = action.get("post", {}).get("date") or action.get("due_window")
    if not date_value:
        return {
            "exit_code": 2,
            "command": str(GBP_POSTER_SCRIPT),
            "stdout": "",
            "stderr": "GBP post action is missing a date.",
        }
    command = ["node", str(GBP_POSTER_SCRIPT), "--date", date_value, "--config", str(GBP_POSTER_CONFIG)]
    if not live:
        command.append("--dry-run")
    if live and GBP_POSTER_HEADLESS:
        command.append("--headless")
    return _run_subprocess(
        command,
        cwd=str(GBP_POSTER_SCRIPT.parent),
        timeout_s=GBP_POSTER_TIMEOUT_S,
    )


def _run_gbp_profile_adapter(action: dict[str, Any], live: bool) -> dict[str, Any]:
    adapter_path = Path(GBP_PROFILE_ADAPTER)
    if not adapter_path.exists():
        return {
            "exit_code": 127,
            "command": str(adapter_path),
            "stdout": "",
            "stderr": f"GBP profile adapter not found: {adapter_path}",
        }
    payload = {"live": live, "action": action}
    command = ["node", str(adapter_path), "--payload", json.dumps(payload)]
    if not live:
        command.append("--dry-run")
    return _run_subprocess(
        command,
        cwd=str(adapter_path.parent),
        timeout_s=GBP_PROFILE_ADAPTER_TIMEOUT_S,
        display_command=" ".join(command[:3]) + " --payload <json>",
    )


def _run_facebook_poster(action: dict[str, Any], live: bool) -> dict[str, Any]:
    adapter_path = Path(FACEBOOK_POSTER_ADAPTER)
    if not adapter_path.exists():
        return {
            "exit_code": 127,
            "command": str(adapter_path),
            "stdout": "",
            "stderr": f"Facebook poster adapter not found: {adapter_path}",
        }
    payload = {"live": live, "action": action}
    command = ["node", str(adapter_path), "--payload", json.dumps(payload)]
    if not live:
        command.append("--dry-run")
    return _run_subprocess(
        command,
        cwd=str(adapter_path.parent),
        timeout_s=FACEBOOK_POSTER_TIMEOUT_S,
        display_command=" ".join(command[:3]) + " --payload <json>",
    )


def facebook_adapter_status() -> dict[str, Any]:
    page_id = os.getenv("FB_PAGE_ID", "").strip()
    access_token = os.getenv("FB_PAGE_ACCESS_TOKEN", "").strip()
    adapter_path = Path(FACEBOOK_POSTER_ADAPTER)
    missing = []
    if not page_id:
        missing.append("FB_PAGE_ID not set in .env")
    if not access_token:
        missing.append("FB_PAGE_ACCESS_TOKEN not set in .env")
    if not adapter_path.exists():
        missing.append("Facebook Graph API adapter script not found")
    state = "live_ready" if (page_id and access_token and adapter_path.exists()) else "blocked"
    return {
        "name": "facebook-graph-api",
        "adapter": str(adapter_path),
        "page_id_set": bool(page_id),
        "access_token_set": bool(access_token),
        "state": state,
        "missing": missing,
    }


def _load_gbp_config() -> dict[str, Any]:
    if not GBP_POSTER_CONFIG.exists():
        raise FileNotFoundError(f"GBP poster config not found: {GBP_POSTER_CONFIG}")
    return json.loads(GBP_POSTER_CONFIG.read_text(encoding="utf-8"))


def gbp_adapter_status() -> dict[str, Any]:
    status: dict[str, Any] = {
        "name": "gbp-browser-poster",
        "script": str(GBP_POSTER_SCRIPT),
        "config": str(GBP_POSTER_CONFIG),
        "state": "missing",
        "browser_session_ready": False,
        "media_upload_ready": False,
        "workbook_ready": False,
        "photo_folder_ready": False,
        "missing": [],
    }
    if not GBP_POSTER_SCRIPT.exists():
        status["missing"].append("GBP poster script")
        return status
    if not GBP_POSTER_CONFIG.exists():
        status["missing"].append("GBP poster config")
        return status
    try:
        config = _load_gbp_config()
    except Exception as error:
        status["state"] = "error"
        status["missing"].append(str(error))
        return status

    status["browser_session_ready"] = GBP_BROWSER_SESSION_DIR.exists()
    status["media_upload_ready"] = all(
        bool(config.get(key))
        for key in ("supabase_url", "supabase_service_role_key", "supabase_bucket")
    )
    status["workbook_ready"] = bool(config.get("workbook_path")) and Path(config["workbook_path"]).exists()
    status["photo_folder_ready"] = bool(config.get("curated_photo_folder")) and Path(config["curated_photo_folder"]).exists()

    if not status["browser_session_ready"]:
        status["missing"].append("Google Business Profile browser session")
    if not status["media_upload_ready"]:
        status["missing"].append("Supabase media upload config")
    if not status["workbook_ready"]:
        status["missing"].append("GBP approval workbook")
    if not status["photo_folder_ready"]:
        status["missing"].append("Curated GBP photo folder")

    if status["browser_session_ready"] and status["workbook_ready"] and status["photo_folder_ready"]:
        status["state"] = "live_ready"
    elif status["workbook_ready"] and status["photo_folder_ready"]:
        status["state"] = "approval_ready"
    else:
        status["state"] = "blocked"
    return status


def _open_gbp_workbook():
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to sync the GBP workbook.") from exc
    config = _load_gbp_config()
    workbook_path = Path(config["workbook_path"])
    if not workbook_path.exists():
        raise FileNotFoundError(f"GBP workbook not found: {workbook_path}")
    workbook = load_workbook(workbook_path)
    if "Posts" not in workbook.sheetnames:
        raise RuntimeError("GBP workbook must contain a Posts sheet.")
    sheet = workbook["Posts"]
    headers = [cell.value for cell in sheet[1]]
    missing = [header for header in GBP_WORKBOOK_HEADERS if header not in headers]
    if missing:
        raise RuntimeError(f"GBP workbook missing headers: {missing}")
    columns = {header: headers.index(header) + 1 for header in GBP_WORKBOOK_HEADERS}
    return config, workbook_path, workbook, sheet, columns


def _save_gbp_workbook(workbook: Any, workbook_path: Path) -> None:
    try:
        workbook.save(workbook_path)
    except PermissionError as exc:
        raise RuntimeError(
            f"Could not save GBP workbook — it is likely open in Excel. Close it and retry: {workbook_path}"
        ) from exc


def _row_date(value: Any) -> str:
    if hasattr(value, "date"):
        return value.date().isoformat()
    if isinstance(value, str):
        return value[:10]
    return ""


def _find_workbook_row(sheet: Any, columns: dict[str, int], date_value: str) -> int | None:
    for row in range(2, sheet.max_row + 1):
        if _row_date(sheet.cell(row, columns["Date"]).value) == date_value:
            return row
    return None


def _caption_for_post(post: dict[str, str]) -> str:
    def clean_gbp_text(value: str) -> str:
        # Google Business Profile rejects post descriptions that include phone numbers.
        value = re.sub(r"\+?1?[\s.-]*\(?\d{3}\)?[\s.-]*\d{3}[\s.-]*\d{4}", "", value)
        value = re.sub(r"\s{2,}", " ", value)
        value = re.sub(r"\s+([.!?,])", r"\1", value)
        return value.strip()

    parts = [
        clean_gbp_text(post.get("headline", "")),
        clean_gbp_text(post.get("body", "")),
        clean_gbp_text(post.get("cta", "")),
    ]
    return "\n\n".join(part for part in parts if part)


def sync_gbp_schedule_to_workbook(dry_run: bool = False) -> dict[str, Any]:
    config, workbook_path, workbook, sheet, columns = _open_gbp_workbook()
    curated_folder = Path(config.get("curated_photo_folder", ""))
    posts = _parse_gbp_post_blocks(_read_text(OUTPUT_DIR / "gbp_posting_schedule.md"))
    updates: list[dict[str, Any]] = []
    next_new_row = sheet.max_row + 1
    for post in posts:
        date_value = post.get("date", "")
        if not date_value:
            continue
        row = _find_workbook_row(sheet, columns, date_value)
        is_new = row is None
        if row is None:
            row = next_new_row
            next_new_row += 1
        photo_file = post.get("photo_file", "")
        photo_path = curated_folder / photo_file if photo_file else Path("")
        updates.append({"date": date_value, "row": row, "new": is_new, "title": post.get("headline", "")})
        if dry_run:
            continue
        existing_status = str(sheet.cell(row, columns["Status"]).value or "").strip()
        existing_posted = bool(sheet.cell(row, columns["Posted"]).value)
        existing_image_link = sheet.cell(row, columns["ImageLink"]).value
        existing_post_url = sheet.cell(row, columns["GBPPostUrl"]).value
        existing_posted_at = sheet.cell(row, columns["PostedAt"]).value
        sheet.cell(row, columns["Date"]).value = date_value
        sheet.cell(row, columns["PostType"]).value = "STANDARD"
        sheet.cell(row, columns["Topic"]).value = post.get("topic") or post.get("service") or post.get("headline")
        sheet.cell(row, columns["AssetSource"]).value = "Workspace Shared"
        sheet.cell(row, columns["AssetIdOrDescription"]).value = str(photo_path) if photo_file else ""
        sheet.cell(row, columns["CTA"]).value = post.get("cta", "")
        sheet.cell(row, columns["Status"]).value = existing_status if existing_status in {"Approved", "Posted"} else post.get("status") or "Needs approval"
        sheet.cell(row, columns["CaptionDraft"]).value = _caption_for_post(post)
        sheet.cell(row, columns["ImageLink"]).value = existing_image_link
        sheet.cell(row, columns["Posted"]).value = existing_posted
        sheet.cell(row, columns["PostedAt"]).value = existing_posted_at
        sheet.cell(row, columns["GBPPostUrl"]).value = existing_post_url
        sheet.cell(row, columns["Notes"]).value = f"Synced from SEO Agents action queue at {_now_iso()}; {post.get('trend_tie', '')}"
    backup_path = None
    if not dry_run:
        backup_path = workbook_path.with_suffix(f".backup-seo-sync-{datetime.now(UTC).strftime('%Y%m%d%H%M%S')}{workbook_path.suffix}")
        shutil.copy2(workbook_path, backup_path)
        _save_gbp_workbook(workbook, workbook_path)
    return {
        "workbook_path": str(workbook_path),
        "backup_path": str(backup_path) if backup_path else None,
        "dry_run": dry_run,
        "posts_found": len(posts),
        "updates": updates,
    }


def mark_gbp_dates_approved(dates: list[str]) -> dict[str, Any]:
    """Stamp the workbook Status to 'Approved' for the given post dates.

    Weekly approval is a single decision: when the week is approved, every day in
    that week should post on its scheduled date. The GBP poster gates on the
    workbook Status column, so this propagates the one weekly approval to all
    days' rows. Rows already 'Posted' are left untouched.
    """
    _, workbook_path, workbook, sheet, columns = _open_gbp_workbook()
    approved: list[str] = []
    skipped: list[str] = []
    for date_value in dates:
        if not date_value:
            continue
        row = _find_workbook_row(sheet, columns, date_value)
        if not row:
            skipped.append(date_value)
            continue
        existing = str(sheet.cell(row, columns["Status"]).value or "").strip()
        if existing == "Posted":
            skipped.append(date_value)
            continue
        sheet.cell(row, columns["Status"]).value = "Approved"
        sheet.cell(row, columns["Notes"]).value = f"Approved (weekly) at {_now_iso()}"
        approved.append(date_value)
    if approved:
        _save_gbp_workbook(workbook, workbook_path)
    return {"workbook_path": str(workbook_path), "approved": approved, "skipped": skipped}


def _mark_gbp_workbook_status(action: dict[str, Any], status: str) -> None:
    date_value = action.get("post", {}).get("date") or action.get("due_window")
    if not date_value:
        return
    _, workbook_path, workbook, sheet, columns = _open_gbp_workbook()
    row = _find_workbook_row(sheet, columns, date_value)
    if not row:
        return
    sheet.cell(row, columns["Status"]).value = status
    sheet.cell(row, columns["Notes"]).value = f"{status} from SEO Agents action queue at {_now_iso()}"
    _save_gbp_workbook(workbook, workbook_path)


def _mark_gbp_workbook_posted(action: dict[str, Any], post_url: str | None = None) -> None:
    date_value = action.get("post", {}).get("date") or action.get("due_window")
    if not date_value:
        return
    _, workbook_path, workbook, sheet, columns = _open_gbp_workbook()
    row = _find_workbook_row(sheet, columns, date_value)
    if not row:
        return
    sheet.cell(row, columns["Status"]).value = "Posted"
    sheet.cell(row, columns["Posted"]).value = True
    sheet.cell(row, columns["PostedAt"]).value = _now_iso()
    if post_url:
        sheet.cell(row, columns["GBPPostUrl"]).value = post_url
    sheet.cell(row, columns["Notes"]).value = "Posted and verified by MCC browser adapter."
    _save_gbp_workbook(workbook, workbook_path)


# ---------------------------------------------------------------------------
# Session 4 — review classification and failure-specific recovery
# ---------------------------------------------------------------------------

def classify_review_failure(action: dict[str, Any]) -> str:
    """Classify an action failure into a recovery category.

    Returns one of:
      - "transient_retry"     — adapter timeout / network blip
      - "evidence_access"     — missing evidence or tool access
      - "contradiction_stall" — unresolved contradiction blocking task
      - "confidence_gap"      — low confidence, needs more research
      - "secrets_quarantine"  — secrets detected in evidence
      - "unknown"
    """
    status = action.get("status", "")
    last_run = action.get("last_run", {})
    if isinstance(last_run, dict):
        cmd_stderr = last_run.get("command_result", {}).get("stderr", "")
    else:
        cmd_stderr = ""
    evidence = action.get("evidence", {})
    if isinstance(evidence, dict):
        evidence = evidence.get("evidence_excerpt", "")
    else:
        evidence = ""
    rejection = action.get("rejection_reason", "")
    confidence = action.get("confidence", {})
    if isinstance(confidence, dict):
        conf_label = confidence.get("label", "")
    else:
        conf_label = ""
    blocker = action.get("blocker", "")
    if isinstance(blocker, str):
        pass
    else:
        blocker = ""

    # Secrets quarantine — highest priority
    if "secret" in cmd_stderr.lower() or "api_key" in evidence.lower() or "credential" in evidence.lower():
        return "secrets_quarantine"
    if "secret" in rejection.lower() or "credential" in rejection.lower():
        return "secrets_quarantine"

    # Evidence access — missing tool access or evidence
    if status == "blocked_access" or "access" in blocker.lower() or "tool" in blocker.lower():
        return "evidence_access"
    if "access" in cmd_stderr.lower() or "tool" in cmd_stderr.lower():
        return "evidence_access"

    # Contradiction stall — unresolved material contradiction
    if status == "blocked" or "contradiction" in blocker.lower() or "contradiction" in rejection.lower():
        return "contradiction_stall"
    if "contradiction" in cmd_stderr.lower():
        return "contradiction_stall"

    # Confidence gap — low confidence needs more research
    if conf_label in ("low", "unknown") and action.get("task_type") != "research_gap":
        return "confidence_gap"

    # Transient retry — adapter timeout or network error
    if "timeout" in cmd_stderr.lower() or "network" in cmd_stderr.lower() or "temporary" in cmd_stderr.lower():
        return "transient_retry"
    if status == "failed" and last_run.get("status") == "adapter_failed":
        return "transient_retry"

    return "unknown"


def apply_recovery(action: dict[str, Any], failure_class: str) -> dict[str, Any]:
    """Apply failure-specific recovery logic and return updated action.

    Each recovery mutates the action dict in place and returns it.
    """
    if failure_class == "transient_retry":
        # Retry once — bump retry count, clear failed status
        retries = action.get("retries", 0) + 1 if action.get("retries") else 1
        action["retries"] = retries
        if action.get("status") == "failed":
            action["status"] = "ready"  # requeue
        action.setdefault("recovery_notes", []).append(f"transient_retry #{retries}")

    elif failure_class == "evidence_access":
        # Escalate — create research_gap task, mark as waiting_on_tool_access
        action["status"] = "waiting_on_tool_access"
        action.setdefault("recovery_notes", []).append("evidence_access_escalation")

    elif failure_class == "contradiction_stall":
        # Escalate — create research_gap task for contradiction resolution
        action["status"] = "blocked"
        action.setdefault("recovery_notes", []).append("contradiction_stall_escalation")

    elif failure_class == "confidence_gap":
        # Create research_gap task for more evidence collection
        action.setdefault("recovery_notes", []).append("confidence_gap_research_task")
        # Don't change status — let the queue keep it pending

    elif failure_class == "secrets_quarantine":
        # Quarantine — remove action from queue, log quarantine
        action["status"] = "quarantined"
        action.setdefault("recovery_notes", []).append("secrets_quarantined")

    return action


# ---------------------------------------------------------------------------
# Session 4 — deterministic idempotency enforcement
# ---------------------------------------------------------------------------

def enforce_idempotency(action: dict[str, Any], live: bool = False) -> dict[str, Any]:
    """Deterministic idempotency guard for website / GBP / Facebook paths.

    When live=True and the action already has a successful run record with the
    same idempotency_key, returns early with the prior result instead of
    re-running the adapter. This prevents duplicate external side effects.

    Returns the run record dict (same shape as run_action returns).
    """
    idem_key = action.get("idempotency_key", "")
    action_id = action.get("id", "unknown")

    if not live:
        # Dry-run always runs — no dedup in dry mode
        return run_action(action_id, live=False)

    # Live mode: check for prior successful run with the same idempotency key
    runs = _load_latest_runs()
    for record in runs.values():
        if record.get("action_id") == action_id and record.get("status") == "live_complete":
            prior_key = record.get("action", {}).get("idempotency_key", "")
            if prior_key == idem_key:
                # Already executed this exact action — return prior result
                return {
                    "id": f"idem_dedupe_{record['id']}",
                    "action_id": action_id,
                    "live": True,
                    "status": "live_complete",
                    "message": f"Idempotency hit — already completed at {record.get('created_at', 'unknown')}",
                    "action": action,
                    "command_result": record.get("command_result"),
                    "created_at": record.get("created_at"),
                    "idempotency_hit": True,
                    "idempotency_key": idem_key,
                }

    # No prior successful run with this key — proceed to execute
    return run_action(action_id, live=True)


def format_action_queue_text(queue: dict[str, Any]) -> str:
    summary = queue["summary"]
    lines = [
        f"Actions: {summary['total']}",
        f"Needs approval: {summary['needs_approval']}",
        f"Approved: {summary['approved']}",
        f"Blocked access: {summary['blocked_access']}",
        f"Dry-run ready: {summary['dry_run_ready']}",
        "",
    ]
    for action in queue["actions"]:
        lines.append(
            f"- {action['id']} [{action['status']}] {action['title']} "
            f"({action['assigned_agent']} / {action['platform']} / {action['risk']})"
        )
    return "\n".join(lines)
